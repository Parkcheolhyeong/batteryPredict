import openpyxl

wb = openpyxl.load_workbook("./K-WATER_total.xlsx")
for sheetName in wb.sheetnames:
	sheet = wb[sheetName]
	print(sheetName)

	for row in range(1, sheet.max_row + 1):
		result = []
		for col in range(1, sheet.max_column + 1):
			result.append(sheet.cell(row, col).value)

		print(result)